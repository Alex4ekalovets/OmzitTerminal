import json
import logging
import math
import os
import shutil
import threading
from copy import copy

import requests

from PyQt6.QtCore import Qt, QSize, pyqtSignal, QRegularExpression
from PyQt6.QtGui import QIcon, QAction, QRegularExpressionValidator, QColor
from PyQt6.QtWidgets import (
    QWidget,
    QGridLayout, QLabel, QTableView,
    QFileDialog, QToolBar, QComboBox, QPushButton, QSplitter, QLineEdit, QMessageBox,
)
import pandas as pd
from openpyxl.reader.excel import load_workbook

from apps.process.dialogs import OperationChoiceWindow, SelectColumnsWindow, SelectNextPrevOperationWindow
from apps.process.models import PandasModel
from apps.process.settings import URL, COLUMNS
from apps.process.utils.cdw_reader import get_specification
from apps.process.utils.graph import draw_graph
from apps.process.utils.parse_operations import df_handler


class OrderModelSelect(QComboBox):
    clicked = pyqtSignal()

    def showPopup(self):
        self.clicked.emit()
        super(OrderModelSelect, self).showPopup()


class ProcessWidget(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.xlsx_file_name = ''
        self.main_window = main_window
        self.layout = QGridLayout()
        self.layout.setSpacing(10)
        self._create_actions()
        self._create_main_widget_toolbar()
        self.displayed_columns = COLUMNS
        self.search_data = None
        self.orders_models = []
        self.is_planned = False
        self.process_is_upload = False
        self.has_fio_doers = []
        self.is_on_change = True
        self.ids_without_doers = []

        self.setLayout(self.layout)
        self.table_process = QTableView()
        self.table_process.setDisabled(True)
        self.table_process.doubleClicked.connect(self.double_click_process_table)
        self._create_process_table_context_menu()

        self.table_details = QTableView()
        self._create_details_table_context_menu()
        self._connect_actions()

        self.initial_data = pd.DataFrame(
            columns=COLUMNS
        )
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        self.set_table_data(self.table_process, self.initial_data)
        self.splitter.addWidget(self.table_process)
        self.splitter.addWidget(self.table_details)

        self.searchbar = QLineEdit()
        self.searchbar.textChanged.connect(self.search)

        order_model_label = QLabel()
        order_model_label.setText('Выберите заказ-модель:')
        self.order_model_select = OrderModelSelect()
        self.order_model_select.setMinimumWidth(160)
        self.order_model_select.addItem("Не выбрано")
        self.order_model_select.clicked.connect(self.on_order_model_click)
        self.order_model_select.activated.connect(self.on_order_model_select)

        model_label = QLabel()
        model_label.setText('Модель:')
        self.model_edit = QLineEdit()
        self.model_edit.setMinimumWidth(160)
        model_reg_ex = QRegularExpression(r"^[\-A-Za-z0-9]+$")
        model_validator = QRegularExpressionValidator(model_reg_ex, self.model_edit)
        self.model_edit.setValidator(model_validator)
        self.model_edit.setToolTip(
            "Имя модели может содержать только цифры и буквы латинского алфавита и знак '-' тире"
        )
        self.model_edit.setDisabled(True)

        self.process_status_label = QLabel()
        self.process_status_label.setText('Статус:')

        self.table_details.hide()
        columns_select_button = QPushButton(text="", parent=self)
        columns_select_button.setIcon(QIcon("static/img/select-columns.svg"))
        columns_select_button.clicked.connect(self.open_select_columns_dialog)

        self.change_button = QPushButton(text="Корректировка", parent=self)
        self.change_button.hide()
        self.change_button.clicked.connect(self.take_on_change)

        self.layout.addWidget(order_model_label, 0, 0, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.order_model_select, 0, 1, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(model_label, 0, 2, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.model_edit, 0, 3, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.process_status_label, 0, 5, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.change_button, 0, 4, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(columns_select_button, 1, 1, 1, 9, Qt.AlignmentFlag.AlignRight)
        self.layout.addWidget(self.searchbar, 1, 0, 1, 2, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.splitter, 2, 0, 1, 10)

    def _create_main_widget_toolbar(self):
        self.widget_tools = QToolBar("Инструменты")
        self.widget_tools.setIconSize(QSize(35, 35))
        self.widget_tools.addAction(self.save_action)
        self.widget_tools.addAction(self.upload_draws_action)
        self.widget_tools.addAction(self.open_xlsx_action)
        self.widget_tools.addAction(self.upload_process_action)
        self.widget_tools.addAction(self.draw_graph_action)
        self.widget_tools.addAction(self.save_to_xlsx_action)
        self.main_window.addToolBar(Qt.ToolBarArea.TopToolBarArea, self.widget_tools)

    def _create_actions(self):
        self.upload_draws_action = QAction(QIcon("static/img/upload.svg"), "Загрузить детали из CDW", self)
        self.open_xlsx_action = QAction(QIcon("static/img/xlsx.svg"), "&Загрузить техпроцесс из XLSX", self)
        self.add_row_up_action = QAction(QIcon("static/img/add-row-up.png"), "&Добавить строку выше", self)
        self.add_row_down_action = QAction(QIcon("static/img/add-row-down.png"), "&Добавить строку ниже", self)
        self.delete_row_action = QAction(QIcon("static/img/delete-row.png"), "&Удалить строку", self)
        self.add_detail_to_process_action = QAction(QIcon("static/img/add-operation.svg"),
                                                    "&Создать операцию для детали",
                                                    self)
        self.add_operation_action = QAction(QIcon("static/img/add-operation.svg"), "&Добавить операцию", self)
        self.upload_process_action = QAction(QIcon("static/img/upload-process.svg"), "&Загрузить техпроцесс", self)
        self.draw_graph_action = QAction(QIcon("static/img/graph.svg"), "&Нарисовать техпроцесс", self)
        self.save_action = QAction(QIcon("static/img/save.svg"), "&Сохранить изменения", self)
        self.save_to_xlsx_action = QAction(QIcon("static/img/to-xlsx.png"), "&Сохранить в XLSX", self)

    def _connect_actions(self):
        self.upload_draws_action.triggered.connect(self.open_cdw_dialog)
        self.open_xlsx_action.triggered.connect(self.open_xlsx_dialog)
        self.add_row_up_action.triggered.connect(self.add_row_up)
        self.add_row_down_action.triggered.connect(self.add_row_down)
        self.delete_row_action.triggered.connect(self.delete_row)
        self.add_detail_to_process_action.triggered.connect(self.add_operation_for_detail)
        self.add_operation_action.triggered.connect(self.add_operation)
        self.upload_process_action.triggered.connect(self.upload_process)
        self.draw_graph_action.triggered.connect(self.draw_graph)
        self.save_action.triggered.connect(self.save_process_to_file)
        self.save_to_xlsx_action.triggered.connect(self.export_to_xlsx)

    @property
    def process_data(self):
        return self.table_process.model().df

    def open_xlsx_dialog(self):
        if self.is_on_change:
            self.xlsx_file_name = QFileDialog.getOpenFileName(
                self,
                caption="Выберите файл",
                directory=r"D:\\",
                filter="Excel Files (*.xlsx);;",
            )[0]
            if self.xlsx_file_name:
                xlsx_file = pd.ExcelFile(self.xlsx_file_name)
                sheet_names = xlsx_file.sheet_names
                dialog = SelectXlsxSheetWindow(self, sheet_names)
                dialog.exec()

                if dialog.selected_sheet:
                    try:
                        if self.has_fio_doers:
                            data = df_handler(
                                pd.read_excel(self.xlsx_file_name, dialog.selected_sheet),
                                start_id=self.table_process.model().last_id
                            )
                            data = pd.concat((self.process_data, data))
                            self.table_process.model().last_id = data[COLUMNS[20]].max() + 1
                        else:
                            data = df_handler(pd.read_excel(self.xlsx_file_name, dialog.selected_sheet))
                            self.table_process.setModel(None)
                        self.set_table_data(self.table_process, data)
                    except Exception as ex:
                        logging.exception(ex)
                else:
                    self.set_table_data(self.table_process, self.initial_data)

    def open_cdw_dialog(self):
        file_names = QFileDialog.getOpenFileNames(
            self,
            caption="Выберите файлы",
            directory=r"D:\\",
            filter="CDW Files (*.cdw);;",
        )[0]
        thread = threading.Thread(target=self.load_cdw, args=(file_names,))
        thread.start()

    def set_table_data(self, table, data):
        logging.debug('Начало построения таблицы')
        data = data.sort_index()
        data.index = range(0, len(data))
        if table.model():
            last_id = table.model().last_id
        else:
            last_id = len(data)
        is_planned_process = table == self.table_process and self.is_planned
        if not self.is_on_change:
            immutables = self.ids_without_doers + self.has_fio_doers
        else:
            immutables = self.has_fio_doers
        model = PandasModel(data, last_id, is_planned_process, immutables=immutables)
        try:
            table.setModel(model)
            logging.debug('Завершение построения таблицы')
        except Exception as ex:
            logging.exception(ex)

    def load_cdw(self, file_names):
        if file_names:
            self.main_window.statusBar().showMessage("Загрузка файлов...")
            json_file = get_specification(file_names)
            data = pd.read_json(json_file, orient='records')
            os.remove(json_file)
            if data.empty:
                data = pd.DataFrame()
            self.set_table_data(self.table_details, data)
            self.main_window.statusBar().clearMessage()
            self.table_details.show()

    def send_tech_data(self, data):
        try:
            self.main_window.statusBar().showMessage("Отправка техпроцесса...")
            url = URL + "tehnolog/tech_data"
            payload = json.dumps(data)
            response = requests.post(url, data=payload)
            data = json.loads(response.content)
            logging.debug(data['message'])
            if response.status_code == 200:
                self.save_process_to_file()
                self.update_process_status()
            self.main_window.statusBar().showMessage(data['message'])
        except requests.exceptions.ConnectionError as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("⛔Нет связи с сервером!")
        except Exception as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("Ошибка отправки техпроцесса!")

    def upload_process(self):
        try:
            data = self.process_data
            if self.is_valid(data):
                shift_tasks = {
                    'model_order_query': f'{self.order_model_select.currentText()}',
                    'order': self.order_model_select.currentText().split('_')[0],
                    'model_name': f'{self.model_edit.text()}',
                    'shift_tasks': [],
                }
                for i in range(len(data)):
                    shift_task = {
                        'model_order_query': f'{self.order_model_select.currentText()}',
                        'order': self.order_model_select.currentText().split('_')[0],
                        'model_name': f'{self.model_edit.text()}',
                        'op_number': f'{data.iloc[i, 0]}',
                        'op_name': f'{data.iloc[i, 1]}',
                        'ws_name': f'{data.iloc[i, 2]}',
                        'op_name_full': f'{data.iloc[i, 1]}-{data.iloc[i, 2]}',
                        'ws_number': f'{data.iloc[i, 3]}',
                        'norm_tech': float(data.iloc[i, 11]),
                        'draw_filename': f'{data.iloc[i, 19]}',
                        'tech_id': int(data.iloc[i, 20]),
                        'next_ids': [int(x) for x in data.iloc[i, 21]],
                        'prev_ids': [int(x) for x in data.iloc[i, 22]],
                    }
                    shift_tasks['shift_tasks'].append(shift_task)
                self.send_tech_data(shift_tasks)
        except Exception as ex:
            logging.exception(ex)

    def is_valid(self, data):
        try:
            message = "⛔"
            terminals = ['11', '12']

            # Проверка 1 - Уникальность id
            is_valid = data[COLUMNS[20]].is_unique
            logging.debug(f"ID уникальны: {data[COLUMNS[20]].is_unique}")

            # Проверка 2 - Все обязательные поля заполнены (столбцы 0, 1, 3, 11)
            empty_cells = data[[COLUMNS[0], COLUMNS[1], COLUMNS[3], COLUMNS[11]]].stack(dropna=False)
            cells = [list(x) for x in empty_cells.index[empty_cells.isna() | empty_cells.isin(['', ])]]
            is_valid = is_valid and not cells
            if cells:
                for cell in cells:
                    row = cell[0]
                    column = COLUMNS.index(cell[1])
                    self.table_process.model().change_color(row, column, QColor("red"))
                message += "Заполните обязательные поля, выделенные красным! "
                is_valid = False

            # Проверка 3 - Номера терминалов указаны верно
            cells = [x for x in data.index[~data[COLUMNS[3]].astype(str).isin(terminals)]]
            if cells:
                for cell in cells:
                    row = cell
                    column = 3
                    self.table_process.model().change_color(row, column, QColor("red"))
                message += f"Номера терминалов должны быть из списка {terminals} "
                is_valid = False

            # Проверка 4 - Заполнена модель
            if self.model_edit.text() == '':
                message += 'Заполните модель! '
                is_valid = False

            # Проверка 5 - Выбран заказ-модель
            if self.order_model_select.currentText() == 'Не выбрано':
                message += 'Выберите заказ-модель! '
                is_valid = False

            self.main_window.statusBar().showMessage(message)
            return is_valid
        except Exception as ex:
            logging.exception(ex)

    def _create_process_table_context_menu(self):
        for action in self.table_process.actions():
            self.table_process.removeAction(action)
        self.table_process.setContextMenuPolicy(Qt.ContextMenuPolicy.ActionsContextMenu)
        self.table_process.addAction(self.add_operation_action)
        self.table_process.addAction(self.add_row_up_action)
        self.table_process.addAction(self.add_row_down_action)
        self.table_process.addAction(self.delete_row_action)

    def _create_details_table_context_menu(self):
        self.table_details.setContextMenuPolicy(Qt.ContextMenuPolicy.ActionsContextMenu)
        self.table_details.addAction(self.add_detail_to_process_action)
        self.table_details.addAction(self.add_row_up_action)
        self.table_details.addAction(self.add_row_down_action)
        self.table_details.addAction(self.delete_row_action)

    def add_empty_row(self, table, name):
        df = table.model().df
        df.loc[name] = ''
        if table == self.table_process:
            table.model().last_id += 1
            df.loc[name, 'id'] = table.model().last_id
            df.at[name, COLUMNS[21]] = []
            df.at[name, COLUMNS[22]] = []
        return df

    def add_row(self, coeff):
        try:
            table = self.focusWidget()
            index = table.currentIndex().row()
            data = table.model().df
            if index == -1 and len(data) == 0:
                data = self.add_empty_row(table, 0)
            else:
                current_row = data.iloc[index].name
                new_name = current_row + coeff
                data = self.add_empty_row(table, new_name)
            self.set_table_data(table, data)
        except Exception as ex:
            logging.exception(ex)

    def add_row_up(self):
        self.add_row(coeff=-0.5)

    def add_row_down(self):
        self.add_row(coeff=0.5)

    def delete_row(self):
        try:
            table = self.focusWidget()
            indexes = set(map(lambda x: x.row(), table.selectedIndexes()))
            data = table.model().df
            row_names = []
            for index in indexes:
                if data.iloc[index, 20] not in self.has_fio_doers:
                    row_names.append(data.iloc[index].name)
            for name in row_names:
                operation_id = data.loc[name, 'id']
                if operation_id not in table.model().immutables:
                    mask = data.iloc[:, 21].apply(lambda x: operation_id in x)
                    data[mask].iloc[:, 21].apply(lambda x: x.remove(operation_id))
                    mask = data.iloc[:, 22].apply(lambda x: operation_id in x)
                    data[mask].iloc[:, 22].apply(lambda x: x.remove(operation_id))
                    data = data.drop(name)
            self.set_table_data(table, data)
        except Exception as ex:
            logging.exception(ex)

    def add_operation_for_detail(self):
        indexes = set(map(lambda x: x.row(), self.table_details.selectedIndexes()))
        data: pd.DataFrame = self.process_data
        dialog = OperationChoiceWindow(self)
        dialog.exec()
        for index in sorted(indexes):
            detail = self.table_details.model().df.loc[index]
            try:
                for operation in dialog.operations:
                    row = len(data)
                    prev_operation_id = data[
                        (data.iloc[:, 15] == detail.iloc[0]) &
                        (data.iloc[:, 16] == detail.iloc[1]) &
                        (data.iloc[:, 17] == detail.iloc[2]) &
                        (data.iloc[:, 18] == detail.iloc[3]) &
                        (data.iloc[:, 19] == detail.iloc[4])
                        ]['id'].max()
                    data = self.add_empty_row(self.table_process, row)
                    data.loc[row, :COLUMNS[15]] = operation
                    data.loc[row, COLUMNS[15]:COLUMNS[19]] = detail
                    if not math.isnan(prev_operation_id):
                        data.loc[row, COLUMNS[22]].append(prev_operation_id)
                        next_operation_id = data.loc[row, 'id']
                        data.loc[data['id'] == prev_operation_id][COLUMNS[21]].iloc[0].append(next_operation_id)

            except Exception as ex:
                logging.exception(ex)
        self.set_table_data(self.table_process, data)

    def add_operation(self):
        try:
            data = self.process_data
            index = self.table_process.currentIndex().row()
            dialog = OperationChoiceWindow(self)
            dialog.exec()
            if index == -1:
                current_row = len(data)
            else:
                current_row = data.iloc[index].name

            for i in range(len(dialog.operations)):
                new_name = current_row + float(f'0.{i + 1}')
                data = self.add_empty_row(self.table_process, new_name)
                data.loc[new_name, :COLUMNS[15]] = dialog.operations[i]
            self.set_table_data(self.table_process, data)
        except Exception as ex:
            logging.exception(ex)

    def open_select_columns_dialog(self):
        dialog = SelectColumnsWindow(self)
        dialog.exec()
        self.displayed_columns = dialog.selected_columns
        for i, value in enumerate(COLUMNS):
            self.table_process.setColumnHidden(i, value not in self.displayed_columns)

    def double_click_process_table(self, index):
        try:
            if index.column() in [21, 22]:
                data = self.process_data
                current_ids = copy(data.iloc[index.row(), index.column()])
                current_id = data.iloc[index.row(), 20]
                filtered_data = data[data.iloc[:, 20] != data.iloc[index.row(), 20]]
                if index.column() == 21:
                    logging.debug('Диалоговое окно выбора следующей операции')
                    column = 22
                elif index.column() == 22:
                    logging.debug('Диалоговое окно выбора предыдущей операции')
                    column = 21
                dialog = SelectNextPrevOperationWindow(self, filtered_data, current_ids)
                dialog.exec()
                data.iloc[index.row(), index.column()].clear()
                data.iloc[index.row(), index.column()].extend(dialog.operations_ids)
                for operation_id in dialog.operations_ids:
                    if current_id not in data.loc[data['id'] == operation_id].iloc[0, column]:
                        data.loc[data['id'] == operation_id].iloc[0, column].append(current_id)
                for operation_id in current_ids:
                    if operation_id not in dialog.operations_ids:
                        data.loc[data['id'] == operation_id].iloc[0, column].remove(current_id)
                self.set_table_data(self.table_process, data)
        except Exception as ex:
            logging.exception(ex)

    def draw_graph(self):
        draw_graph(self.process_data)

    def search(self):
        try:
            if self.search_data is None:
                self.search_data = self.process_data
            data = self.search_data
            data = data[
                data.apply(lambda x: x.astype(str).str.contains(self.searchbar.text(), case=False, regex=False)).any(
                    axis=1)]
            self.set_table_data(self.table_process, data)
        except Exception as ex:
            logging.exception(ex)

    def get_orders_models(self):
        try:
            url = URL + "tehnolog/orders_models"
            response = requests.get(url)
            self.orders_models = json.loads(response.content)
            self.main_window.statusBar().clearMessage()
        except requests.exceptions.ConnectionError as ex:
            logging.exception(ex)
            self.not_connection()
            self.main_window.statusBar().showMessage("⛔Нет связи с сервером!")
        except Exception as ex:
            logging.exception(ex)
            self.not_connection()
            self.main_window.statusBar().showMessage("Ошибка получения данных!")

    def not_connection(self):
        try:
            self.orders_models = []
            self.model_edit.setDisabled(True)
            self.model_edit.clear()
            self.table_process.setDisabled(True)
        except Exception as ex:
            logging.exception(ex)

    def on_order_model_click(self):
        try:
            self.order_model_select.clear()
            self.order_model_select.addItem("Не выбрано")
            self.get_orders_models()
            for order_model in self.orders_models:
                self.order_model_select.addItem(order_model['order_model'])
        except Exception as ex:
            logging.exception(ex)

    def save_process_to_file(self):
        try:
            order = self.order_model_select.currentText().split('_')[0]
            model = self.model_edit.text()
            order_model = f'{order}_{model}'

            if len(self.searchbar.text()) != 0:
                data = self.search_data
                logging.debug('Примененный фильтр сброшен перед сохранением')
            else:
                data = self.process_data

            file = fr'data\process\{order_model}.json'
            if data.empty:
                try:
                    os.remove(file)
                    logging.debug(f'Сохранен пустой процесс. Файл {file} удален')
                except Exception as ex:
                    logging.debug(f'При попытке удаления файла {file} вызвано исключение: {ex}')
            else:
                data.to_json(file, orient='records')
                logging.debug(fr'Процесс сохранен в файл {file}')
                self.main_window.statusBar().showMessage("💾Сохранено!")
        except Exception as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("⛔Ошибка сохранения!")

    def open_process_from_file(self, order_model):
        try:
            file = fr'data\process\{order_model["order_model"]}.json'
            if os.path.exists(file):
                data = pd.read_json(file, orient='records', dtype={
                    COLUMNS[i]: object for i in list(range(6)) + list(range(15, 20))
                })
                logging.debug(fr'Открыт файл процесса {file}')
            elif order_model['td_status'] == 'утверждено':
                data = self.create_rows_from_shift_tasks(order_model["order_model"])
            else:
                logging.debug('Новый процесс')
                data = self.initial_data

            max_id = data[COLUMNS[20]].max()
            self.table_process.model().last_id = max_id if pd.notna(max_id) else 0
            self.set_table_data(self.table_process, data)
        except Exception as ex:
            logging.exception(ex)

    def export_to_xlsx(self):
        try:
            order = self.order_model_select.currentText().split('_')[0]
            model = self.model_edit.text()
            order_model = f'{order}_{model}'
            name = QFileDialog.getSaveFileName(self, 'Сохранить файл', f"{order_model}.xlsx")[0]
            if name:
                shutil.copyfile('templates/process/export_process_template.xlsx', name)
                new_rows_count = len(self.process_data)
                wb = load_workbook(name)
                ws = wb['Sheet1']
                ws.title = model
                ws['B2'] = f"Расценки на работу по котлу  {model}"
                ws.insert_rows(5, amount=new_rows_count)
                ws[f'L{5 + new_rows_count}'] = f'=SUM(L5:L{4 + new_rows_count})'
                ws[f'M{5 + new_rows_count}'] = f'=SUM(M5:M{4 + new_rows_count})'
                ws[f'O{5 + new_rows_count}'] = f'=SUM(O5:O{4 + new_rows_count})'
                ws[f'L{7 + new_rows_count}'] = f'=L{5 + new_rows_count}/G2'
                wb.save(name)
                with pd.ExcelWriter(name, if_sheet_exists='overlay', mode='a') as writer:
                    self.process_data.to_excel(
                        excel_writer=writer,
                        sheet_name=model,
                        index=False,
                        header=False,
                        float_format="%.2f",
                        columns=COLUMNS[:15],
                        startrow=4,
                    )
                logging.debug(f'Процесс сохранен в файл Excel {name}')
        except Exception as ex:
            logging.exception(ex)

    def on_order_model_select(self, item):
        try:
            selected_order_model = self.order_model_select.itemText(item)
            for order_model in self.orders_models:
                if selected_order_model == order_model['order_model']:

                    self.is_planned = order_model['order_status'] != 'не запланировано'
                    self.has_fio_doers = order_model['has_fio_doers']

                    conditions = [
                        order_model['td_status'] == 'утверждено',
                        not order_model['on_change']
                    ]

                    self.process_is_upload = all(conditions)

                    self.process_status_label.setText(
                        f'Статус: {"загружен" if self.process_is_upload else "не загружен"}, '
                        f'{order_model["order_status"]}, '
                        f'{"распределено" if self.has_fio_doers else "не распределено"}'
                    )

                    self.ids_without_doers = order_model["st_without_doers"]

                    if self.process_is_upload and len(self.ids_without_doers) > 0:
                        self.change_button.show()
                        self.is_on_change = False
                    else:
                        self.change_button.hide()
                        self.is_on_change = True

                    self._create_process_table_context_menu()
                    self.model_edit.setText(order_model['model'])
                    self.model_edit.setDisabled(self.is_planned)
                    self.table_process.setDisabled(False)

                    self.open_process_from_file(order_model)
                    break
                else:
                    self.change_button.hide()
                    self.model_edit.clear()
                    self.model_edit.setDisabled(True)
                    self.table_process.setDisabled(True)
                    self.process_status_label.setText('Статус: ')
        except Exception as ex:
            logging.exception(ex)

    def update_process_status(self):
        order = self.order_model_select.currentText().split('_')[0]
        self.on_order_model_click()
        self.order_model_select.setCurrentText(f'{order}_{self.model_edit.text()}')
        self.on_order_model_select(self.order_model_select.currentIndex())

    def set_change_status(self, data):
        try:
            self.main_window.statusBar().showMessage("Изменение статуса...")
            url = URL + "tehnolog/change_st_status"
            payload = json.dumps(data)
            response = requests.post(url, data=payload)
            data = json.loads(response.content)
            if response.status_code == 200:
                self.update_process_status()
            self.main_window.statusBar().showMessage(data['message'])
            logging.debug(data['message'])
        except requests.exceptions.ConnectionError as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("⛔Нет связи с сервером!")
        except Exception as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("Ошибка отправки запроса!")

    def take_on_change(self):
        try:
            message = (f'Вы уверены, что хотите откорректировать процесс?\n'
                       f'При корректировке сменные задания будут недоступны\n'
                       f'для планирования и распределения.\n'
                       f'Для возможности дальнейшего распределения необходимо\n'
                       f'обязательно повторно загрузить процесс!')
            answer = QMessageBox.critical(
                self,
                'Предупреждение!',
                message,
                buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel)
            if answer == QMessageBox.StandardButton.Yes:
                ids = self.process_data[COLUMNS[20]].tolist()
                data = {
                    'model_order_query': self.order_model_select.currentText(),
                    'tech_ids': list(set(ids) - set(self.has_fio_doers)),
                    'status': 'корректировка'
                }
                self.set_change_status(data)
        except Exception as ex:
            logging.exception(ex)

    def create_rows_from_shift_tasks(self, order_model):
        shift_tasks = self.get_shift_tasks(order_model)
        data = []
        for st in shift_tasks:
            data.append(
                {
                    'п/п': st['op_number'],  # 0
                    'Наименование работ': st['op_name'],  # 1
                    'Рабочий центр': st['ws_name'],  # 2
                    '№ рабочего центра': st['ws_number'],  # 3
                    'ГОСТ и тип сварочного шва': '',  # 4
                    'ед.изм.': '',  # 5
                    'Объём работ (максимальный в смену)': '',  # 6
                    'Трудоёмкость в смену, час': '',  # 7
                    'Численность, чел.': '',  # 8
                    'Трудоёмкость на 1 ед./заготовку, чел/час': '',  # 9
                    'Кол-во ед./ заготовок на 1 котел': '',  # 10
                    'Трудоёмкость на 1 котел, чел/час': st['norm_tech'],  # 11
                    'Расценка за объем работ, руб.': '',  # 12
                    'Стоимость часа, руб': '',  # 13
                    'Загрузка оборудования на 1 котел, часов': '',  # 14
                    'Наименование': '',  # 15
                    'Материал': '',  # 16
                    'Количество': '',  # 17
                    'Длина': '',  # 18
                    'Чертеж': st['draw_filename'],  # 19
                    'id': st['tech_id'],  # 20
                    'Следующая операция': [],  # 21
                    'Предыдущая операция': []  # 22
                })
        data = pd.DataFrame(data)
        return data

    def get_shift_tasks(self, order_model):
        try:
            self.main_window.statusBar().showMessage("Получение сменных заданий...")
            url = URL + f"tehnolog/shift_tasks/{order_model}"
            response = requests.get(url)
            data = json.loads(response.content)
            if response.status_code == 200:
                self.main_window.statusBar().showMessage('Завершено!')
                return data
        except requests.exceptions.ConnectionError as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("⛔Нет связи с сервером!")
        except Exception as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("Ошибка отправки запроса!")

