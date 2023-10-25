import copy
from typing import List, Dict

import openpyxl
import re
# from ..models import TechData, ProductCategory, ProductModel
from ..models import TechData, ProductCategory
from scheduler.models import ShiftTask


def tech_data_get(model_order_query: str, exel_file: str, excel_list: str):
    """
    Функция получает технологические данные из файла exel_file по списку моделей изделия model_list,
    который равен списку имен листов файла исключая имена листов в списке exclusion_list и сохраняет в модели
    django ProductModel, TechData
    :param model_order_query: заказ-модель
    :param exel_file: имя файла excel
    :param excel_list: модель изделия для получения данных - лист книги excel
    :return: None
    """
    # имя модели в базу
    model_name = model_order_query[model_order_query.find("_") + 1:]
    # номер заказа в базу модели в базу
    order = model_order_query[:model_order_query.find("_")]

    # Получение списка данных
    common_data = {
        'model_name': model_name,
        'order': order,
        'model_order_query': model_order_query
    }
    data_list = get_excel_data(common_data, exel_file, excel_list)

    shift_tasks = ShiftTask.objects.filter(model_order_query=model_order_query)

    def create_all_shift_tasks():
        tasks = [ShiftTask(**data) for data in data_list]
        ShiftTask.objects.bulk_create(tasks)

    is_uploaded = True
    if not shift_tasks.exists():
        create_all_shift_tasks()
    elif all(st == 'не запланировано' for st in shift_tasks.values_list('st_status', flat=True)):
        shift_tasks.delete()
        create_all_shift_tasks()
    else:
        tasks = []
        allowed_fields = ('ws_number', 'norm_tech', 'draw_filename')
        for data in data_list:
            allowed_data = dict()
            for field in allowed_fields:
                allowed_data[field] = data.pop(field)
            try:
                task = ShiftTask.objects.get(**data)
                for key, value in allowed_data.items():
                    setattr(task, key, value)
                tasks.append(task)
            except ShiftTask.DoesNotExist:
                is_uploaded = False
    return is_uploaded




        # changed_count = len(data_list) - len(shift_tasks)
        # found_new_lines_count = 0
        # for data in data_list:
        #     input_row = data['excel_list_name']
        #     updated_data = dict()
        #     for field in allowed_fields:
        #         updated_data[field] = data.pop(field)
        #     try:
        #
        #         if changed_count > 0:
        #             for add_row in range(1, changed_count + 2):
        #                 try:
        #                     task = ShiftTask.objects.get(**data)
        #                     shift_tasks = shift_tasks.exclude(excel_list_name=data['excel_list_name'])
        #                     break
        #                 except ShiftTask.DoesNotExist:
        #                     next_row_number = int(data['excel_list_name'].split('-')[1]) - 1
        #                     data['excel_list_name'] = '-'.join((excel_list, str(next_row_number)))
        #             else:
        #                 found_new_lines_count += 1
        #                 if found_new_lines_count <= changed_count:
        #                     data['excel_list_name'] = input_row
        #                     row = int(data['excel_list_name'].split('-')[-1]) + 1
        #                     data.update(updated_data)
        #                     errors['added_rows'].update({row: data})
        #                 else:
        #                     next_row_number = int(data['excel_list_name'].split('-')[1]) + changed_count - 1
        #                     data['excel_list_name'] = '-'.join((excel_list, str(next_row_number)))
        #                     shift_tasks = shift_tasks.exclude(excel_list_name=data['excel_list_name'])
        #                     raise ShiftTask.DoesNotExist
        #
        #         elif changed_count < 0:
        #             for add_row in range(1, abs(changed_count) + 2):
        #                 try:
        #                     task = ShiftTask.objects.get(**data)
        #                     shift_tasks = shift_tasks.exclude(excel_list_name=data['excel_list_name'])
        #                     break
        #                 except ShiftTask.DoesNotExist:
        #                     next_row_number = int(data['excel_list_name'].split('-')[1]) + 1
        #                     data['excel_list_name'] = '-'.join((excel_list, str(next_row_number)))
        #             else:
        #                 next_row_number = int(data['excel_list_name'].split('-')[1]) - abs(changed_count)
        #                 data['excel_list_name'] = '-'.join((excel_list, str(next_row_number)))
        #                 shift_tasks = shift_tasks.exclude(excel_list_name=data['excel_list_name'])
        #                 raise ShiftTask.DoesNotExist
        #
        #         else:
        #             shift_tasks = shift_tasks.exclude(excel_list_name=data['excel_list_name'])
        #             task = ShiftTask.objects.get(**data)
        #             for key, value in updated_data.items():
        #                 setattr(task, key, value)
        #             tasks.append(task)
        #
        #     except ShiftTask.DoesNotExist:
        #         current_data = ShiftTask.objects.filter(excel_list_name=data['excel_list_name'])
        #         row = int(data['excel_list_name'].split('-')[-1]) + 1
        #         if current_data.exists():
        #             errors['changed_rows'].update({row: (current_data[0], data)})
        #         else:
        #             row = int(data['excel_list_name'].split('-')[-1]) + 1
        #             data.update(updated_data)
        #             errors['added_rows'].update({row: data})
        #
        # for task in shift_tasks:
        #     row = int(task.excel_list_name.split('-')[-1]) + 1
        #     errors['deleted_rows'].update({row: task})
        #
        # ShiftTask.objects.bulk_update(tasks, allowed_fields)
        # return errors if any(len(value) > 0 for key, value in errors.items()) else None


def get_excel_data(data: Dict, exel_file: str, excel_list: str) -> List:
    ex_wb = openpyxl.load_workbook(exel_file, data_only=True)
    excel_list = excel_list.strip()
    ex_sh = ex_wb[excel_list.strip()]

    # шаблон номера операции
    op_number_template = r'\d\d.\d\d*'

    # определение максимальной строки для чтения
    max_read_row = 1
    for row in ex_sh.iter_rows(min_row=2, min_col=1, max_row=ex_sh.max_row,
                               max_col=2, values_only=True):
        max_read_row += 1
        if 'итого' in str(row[1]).lower():
            break

    data_list = []
    for i, row in enumerate(ex_sh.iter_rows(min_row=1, min_col=1, max_row=max_read_row,
                                            max_col=ex_sh.max_column, values_only=True)):
        if re.fullmatch(op_number_template, str(row[0])):
            data.update(
                {
                    'excel_list_name': f'{excel_list}-{i}',
                    'op_number': row[0],
                    'op_name': row[1],
                    'ws_name': row[2],
                    'op_name_full': row[1] + '-' + row[2],
                    'ws_number': str(row[3]),
                    'norm_tech': row[11],
                    'draw_filename': row[15],
                }
            )
            previous_op_number = str(row[0])  # обработка объединенных ячеек имени операции
            previous_op_name = row[1]
        elif row[0] is None and row[2] is not None:  # если объединённая ячейка
            data.update(
                {
                    'excel_list_name': f'{excel_list}-{i}',
                    'op_number': previous_op_number,
                    'op_name': previous_op_name,
                    'ws_name': row[2],
                    'op_name_full': previous_op_name + '-' + row[2],
                    'ws_number': str(row[3]),
                    'norm_tech': row[11],
                    'draw_filename': row[15],
                }
            )
        else:
            continue
        data_list.append(copy.copy(data))
    return data_list


if __name__ == '__main__':
    ex_file_dir_tst = r'D:\АСУП\Python\Projects\OmzitTerminal\Трудоёмкость серия I.xlsx'
    model_list_tst = ['7000М+', '800М+']
    exclusion_list_tst = ('Интерполяция М', 'гофрирование', 'Интерполяция R')
    # tech_data_get(exel_file=ex_file_dir_tst)
